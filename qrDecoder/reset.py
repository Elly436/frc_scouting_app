from openpyxl import load_workbook

workbook = load_workbook(filename="testMatches.xlsx")

for sheet in workbook.sheetnames:
    workbook.remove(workbook[sheet])

workbook.create_sheet("personal_score ranking")

workbook.save(filename="testMatches.xlsx")
