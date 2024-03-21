from openpyxl import load_workbook
import spreadsheetCom as ssc

original_name = "idahotest"
merge_name = "idahotest2"

workbook_original = load_workbook(filename="idahotest.xlsx")
workbook_merge = load_workbook(filename="idahotest2.xlsx")

sheets_to_ignore = ["personal score ranking", "amp ranking", "speaker ranking"]


for sheet in workbook_merge.sheetnames:
    merge_sheet = workbook_merge[sheet]
    if not sheet in workbook_original.sheetnames:
        ssc.createTeamSheet(sheet)
    current_sheet = workbook_original[sheet]

    if not sheet in sheets_to_ignore:

        done = False
        i = 2
        for merge_row in merge_sheet.iter_rows(min_row=2, values_only=True):
            for row in current_sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    print("adding")
                    current_sheet.insert_rows(idx=i)
                    ssc.writeRow(i, list(merge_row), sheet)
                    done = True
                    break

                elif row[0] == merge_row[0]:
                    print("duplicate match for the same team?")
                    done = True
                    break

                elif row[0] > merge_row[0]:
                    print(list(merge_row))
                    current_sheet.insert_rows(idx=i)
                    ssc.writeRow(i, list(merge_row), sheet)
                    done = True
                    break

                i += 1

            i = 2

            if not done:
                print(i)
                print("adding but other other")
                current_sheet.insert_rows(idx=i)
                ssc.writeRow(i, merge_row, sheet)
                    
    print("done " + sheet)