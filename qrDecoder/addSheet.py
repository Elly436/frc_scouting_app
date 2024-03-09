from openpyxl import load_workbook

workbook = load_workbook(filename="testMatches.xlsx")

workbook.create_sheet("score ranking")