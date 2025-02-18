from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

file_name = "idahotest"

categories = ["personal score", "total score", "dcsnl", "dskd", "hullo"]
#categories = ["personal_score", "total_score", "fadjj"]

def getLetter(number):
    charstr = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    chars = list(charstr)
    output = chars[number]
    return output

rankingCategories = ["personal score", "dcsnl", "dskd + hullo"]
rankingsIndex = [[0], [2], [3, 4]] #index of the category it corresponds to

averageCells = [] # the row of the reference for the ranking sheets
for i in range(len(rankingCategories)):
    averageCells.append(getLetter(len(categories) + 5 + i))

averageLast5Cells = []
for i in range(len(rankingCategories)):
    averageCells.append(getLetter(len(categories) + 5 + len(averageCells) + i))

averageCellText = [] # generate the text that makes the averages
for i in rankingsIndex:
    num = 0
    text = "="
    for x in i:
        if num > 0:
            text += "+"
        text += 'AVERAGE(' + getLetter(x) + '2:INDIRECT("'+ getLetter(x)+'"&COUNTA(' + getLetter(x) + ':' + getLetter(x) + ')))'
        num += 1
    averageCellText.append(text)

print(averageCellText)


averageLast5CellText = []
for i in rankingsIndex:
    num = 0
    text = "="
    for x in i:
        if num > 0:
            text += "+"
        text += 'AVERAGE(INDIRECT("'+ getLetter(x)+'"&(COUNTA(' + getLetter(x) + ':' + getLetter(x) + ')-5)):INDIRECT("'+ getLetter(x)+'"&COUNTA(' + getLetter(x) + ':' + getLetter(x) + ')))'
        num += 1
    averageLast5CellText.append(text)

print(averageLast5CellText)

workbook = load_workbook(filename=file_name + ".xlsx")
#sheet = workbook["6364"]


# for category in categories:
#     if not category + " ranking" in workbook.sheetnames:
#         workbook.create_sheet(category + " ranking")

def splitString(toSplit, splitter):
    returnList = [""]
    for i in range(len(toSplit)):
        if toSplit[i] == splitter and not returnList[len(returnList)-1] == "":
            returnList.append("")
        else:
            returnList[len(returnList) - 1] += toSplit[i]
    # for i in range(len(returnList)):
    #     returnList[i] = int(returnList[i])
    return returnList

def writeRow(rowNumber, list, team):
    activeSheet = workbook[team]
    for i in range(len(list)):
        activeSheet[getLetter(i) + str(rowNumber)] = list[i]

def createTeamSheet(team):
    workbook.create_sheet(team)
    
    headers = categories
    headers.append("comment")
    headers.append("outlier")
    headers.insert(0, "match")
    headers.extend([None, "average:"])
    headers.extend(averageCellText)
    headers.extend(["avg last 5:"])
    headers.extend(averageLast5CellText)
    writeRow(1, headers, team)
    workbook[team].freeze_panes = "A2"
    # categories.remove("comment")
    # categories.remove("outlier")
    # categories.remove("match")
    # for thing in [None, "average:", '=INDIRECT("B"&COUNTA(A:A))', '=INDIRECT("I"&COUNTA(A:A))', '=INDIRECT("L"&COUNTA(A:A))', 'avg last 5:', '=INDIRECT("B"&COUNTA(A:A)+1)', '=INDIRECT("I"&COUNTA(A:A)+1)', '=INDIRECT("L"&COUNTA(A:A)+1)']:
    #     categories.remove(thing)

    red_background = PatternFill(bgColor="00FFFF00")
    diff_style = DifferentialStyle(fill= red_background)
    rule = Rule(type="expression", dxf= diff_style)
    rule.formula = ["$P1=1"]
    workbook[team].conditional_formatting.add("A1:O100", rule)

    

    #=AVERAGE(INDIRECT("B"&(COUNTA(B2:B2)-5),TRUE):INDIRECT("B"&(COUNTA(B2:B2)+1),TRUE))
    for i in range(len(rankingCategories)):
        workbook[rankingCategories[i]+" ranking"].append([team, "='"+str(team)+"'!"+averageCells[i]+"1", "='"+str(team)+"'!"+averageLast5Cells[i]+"1"])


def addMatch(string_data):
    data = splitString(string_data, "*")
    data_dict = {}
    for thing in data:
        split_thing = splitString(thing, ":")
        category = split_thing[0]
        value = split_thing[1]
        data_dict[category] = value
    team = str(data_dict.pop("team"))

    if not team in workbook.sheetnames:
        createTeamSheet(team)
    activeSheet = workbook[team]
    matchExists = False

    #print(data_dict)

    i = 2
    for row in activeSheet.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            print("adding match")
            activeSheet.insert_rows(idx=i)
            row_data = []
            for category in categories:
                cell = data_dict[category]
                if cell.isnumeric():
                    cell = int(cell)
                row_data.append(cell)
            row_data.append(data_dict["comment"])
            row_data.append(int(data_dict["outlier"]))
            row_data.insert(0, int(data_dict["match"]))
            writeRow(i, row_data, team)
            matchExists = True
            break
        elif row[0] == int(data_dict["match"]):
            print("override")
            row_data = []
            for category in categories:
                cell = data_dict[category]
                if cell.isnumeric():
                    cell = int(cell)
                row_data.append(cell)
            row_data.append(data_dict["comment"])
            row_data.append(int(data_dict["outlier"]))
            row_data.insert(0, int(data_dict["match"]))
            writeRow(i, row_data, team)
            matchExists = True
            break
        elif row[0] > int(data_dict["match"]):
            print("placing match")
            activeSheet.insert_rows(idx=i)
            row_data = []
            for category in categories:
                cell = data_dict[category]
                if cell.isnumeric():
                    cell = int(cell)
                row_data.append(cell)
            row_data.append(data_dict["comment"])
            row_data.append(int(data_dict["outlier"]))
            row_data.insert(0, int(data_dict["match"]))
            writeRow(i, row_data, team)
            matchExists = True
            break

        i += 1
    if not matchExists:
        print("adding data")
        row_data = []
        for category in categories:
            cell = data_dict[category]
            if cell.isnumeric():
                cell = int(cell)
            row_data.append(cell)
        row_data.append(data_dict["comment"])
        row_data.append(int(data_dict["outlier"]))
        row_data.insert(0, int(data_dict["match"]))
        #print(row_data)
        activeSheet.append(row_data)
    #chart = LineChart()
    # points = Reference(worksheet=activeSheet,
    #                  min_row=1,
    #                  min_col=2,
    #                  max_col=3)
    #chart.add_data(points, from_rows=False, titles_from_data=True)
    #activeSheet.add_chart(chart, getLetter(len(data) + 4)+"2")
    workbook.save(filename=file_name + ".xlsx")

# Using the values_only because you want to return the cells' values
def printRows():
    for row in workbook["6364"].iter_rows(values_only=True):
        print(row[0])


#writeRow(1, ["personal score", "total score", "objects scored"], "score ranking")
#sheet.freeze_panels = "A2"
#printRows()
# Save the spreadsheet
workbook.save(filename=file_name + ".xlsx")
