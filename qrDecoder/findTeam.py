from openpyxl import load_workbook

file_name = "qualifications"

workbook = load_workbook(filename=file_name + ".xlsx"),

team = "4627q"

sheet = workbook[team]

for row in sheet.iter_rows(values_only=True):
    print(row)